# -*- coding: utf-8 -*-
"""
Created on Tue Dec  6 16:48:41 2022

@author: fabic

"""


# Import required libraries:
import logging as log
import xlwt
import numpy as np
import scipy.io
import pandas as pd
import matplotlib.pyplot as plt   
from copy import deepcopy
from tqdm import tqdm
from scipy.interpolate import interp1d
import openpyxl


# Read data into pandas dataframe:
DF = pd.read_excel('shape-internal_snapshot_1670341103.xlsx')
DF.head(5)

# Define region classification
Regions_I  = ['Brazil','Canada','Central Europe','China region','Eastern Africa',                     #0-4
              'India','Indonesia','Japan','Korea','Middle East',                                      #5-9
              'Mexico','Northern Africa','Oceania','Rest of Central America','Rest of South Africa',  #10-14
              'Rest of South America','Rest of South Asia','Russia','South Africa','Southeast Asia',  #15-19
              'Kazakhstan region','Turkey','Ukraine region','USA','Western Africa'                    #20-24
              'Western Europe']                                                                       #25
                                                                   
Model_I    = ['IMAGE 3.3']
Regions_I = [Model_I[0]+'|'+region_i for region_i in Regions_I]

Regions_O = ['World']
Regions_R  = ['R32CAN','R32CHN','R32EU12-M','R32IND','R32JPN','R32USA','France','Germany','Italy','Poland','Spain','UK','Oth_R32EU15','Oth_R32EU12-H','R5.2OECD_Other','R5.2REF_Other','R5.2ASIA_Other','R5.2MNF_Other','R5.2LAM_Other','R5.2SSA_Other']

R_DictLookUp = {
    'R32CAN':         ['Canada'] ,
    'R32CHN':         ['China region'],
    'R32EU12-M':      ['Central Europe'],
    'R32IND':         ['India'],
    'R32JPN':         ['Japan'],
    'R32USA':         ['USA'],
    'France':         ['Western Europe'],
    'Germany':        ['Western Europe'],
    'Italy':          ['Western Europe'],
    'Poland':         ['Central Europe'],
    'Spain':          ['Western Europe'],
    'UK':             ['Western Europe'],
    'Oth_R32EU15':    ['Western Europe'],
    'Oth_R32EU12-H':  ['Central Europe'],
    'R5.2OECD_Other': ['Oceania', 'Mexico','Korea','Turkey'],
    'R5.2REF_Other':  ['Russia','Kazakhstan region','Ukraine region'],
    'R5.2ASIA_Other': ['Indonesia','Rest of South Asia','Southeast Asia'],
    'R5.2MNF_Other':  ['Middle East','Northern Africa'],
    'R5.2LAM_Other':  ['Brazil','Rest of Central America','Rest of South America'],
    'R5.2SSA_Other':  ['South Africa','Rest of South Africa','Western Africa']
   }

#merge model and region name
for key in R_DictLookUp.keys():
    R_DictLookUp[key] = [Model_I[0]+'|'+ r for r in R_DictLookUp[key]]


O_DictLookUp = {'World':['Brazil','Canada','Central Europe','China region','Eastern Africa',                     
              'India','Indonesia','Japan','Korea','Middle East',                                      
              'Mexico','Northern Africa','Oceania','Rest of Central America','Rest of South Africa',  
              'Rest of South America','Rest of South Asia','Russia','South Africa','Southeast Asia',  
              'Kazakhstan region','Turkey','Ukraine region','USA','Western Africa'  ,                  
              'Western Europe'] }

for key in O_DictLookUp.keys():
    O_DictLookUp[key] = [Model_I[0]+'|'+ r for r in O_DictLookUp[key]]
  


# Define scenario classification
Scenario_I = ['SSP2','SDP_EI-1p5C','SDP_MC-1p5C','SDP_RC-1p5C']
RCPScen_R  = ['Baseline(unmitigated)','RCP2.6']

S_DictLookUp = {
    'Baseline(unmitigated)': ['SSP2'],
    'RCP2.6': ['SDP_EI-1p5C']
    } 

# Define time classification
Years_I    = [2005,2010,2015,2020,2025,2030,2035,2040,2045,2050,2060,2070,2080,2090,2100]
Time_R     = np.arange(2005,2101,1)
TimeL_R    = [i for i in Time_R] 


# Define electricity genreation classification
Industry_I = np.array(['Secondary Energy|Electricity|Biomass|w/ CCS',  #0
         'Secondary Energy|Electricity|Biomass|w/o CCS', #1
         'Secondary Energy|Electricity|Coal|w/ CCS',     #2
         'Secondary Energy|Electricity|Coal|w/o CCS',    #3
         'Secondary Energy|Electricity|Gas|w/ CCS',      #4
         'Secondary Energy|Electricity|Gas|w/o CCS',     #5
         'Secondary Energy|Electricity|Geothermal',      #6
         'Secondary Energy|Electricity|Hydro',           #7
         'Secondary Energy|Electricity|Nuclear',         #8
         'Secondary Energy|Electricity|Oil|w/ CCS',      #9
         'Secondary Energy|Electricity|Oil|w/o CCS',     #10
         'Secondary Energy|Electricity|Solar|CSP',       #11
         'Secondary Energy|Electricity|Solar|PV',        #12
         'Secondary Energy|Electricity|Wind|Offshore',   #13
         'Secondary Energy|Electricity|Wind|Onshore'     #14
         ]) # list 15 IMAGE electricity industry, in EJ/yr
         
Industry_R = np.array(['solar photovoltaic power plant',        #0
         'concentrating solar power plant (CSP)', #1
         'wind power plant onshore',              #2
         'wind power plant offshore',             #3
         'hydro power plant',                     #4
         'nuclear power plant',                   #5
         'coal power plant',                      #6
         'coal power plant without abatement measures',#7
         'bio powerplant',                        #8
         'oil power plant',                       #9
         'geothermal power plant',                #10
         'IGCC power plant',                      #11
         'light oil combined cycle',              #12
         'gas combined cycle power plant',        #13
         'advanced coal power plant with CCS',    #14
         'coal power plant with CCS',             #15
         'biomass power plant with CCS',          #16
         'gas combined cycle power plant with CCS'#17
         ]) # list of 18 electricity industry in RECC

Ind_DictLookUp   = {
             'solar photovoltaic power plant':          ['Secondary Energy|Electricity|Solar|PV'],      
             'concentrating solar power plant (CSP)':   ['Secondary Energy|Electricity|Solar|CSP'],
             'wind power plant onshore':                ['Secondary Energy|Electricity|Wind|Onshore'],             
             'wind power plant offshore':               ['Secondary Energy|Electricity|Wind|Offshore'],             
             'hydro power plant':                       ['Secondary Energy|Electricity|Hydro'],                     
             'nuclear power plant':                     ['Secondary Energy|Electricity|Nuclear'],                   
             'coal power plant':                        ['Secondary Energy|Electricity|Coal|w/o CCS'],                      
             'coal power plant without abatement measures': [],
             'bio powerplant':                          ['Secondary Energy|Electricity|Biomass|w/o CCS'],                        
             'oil power plant':                         ['Secondary Energy|Electricity|Oil|w/ CCS', 
                                                         'Secondary Energy|Electricity|Oil|w/o CCS'],                       
             'geothermal power plant':                  ['Secondary Energy|Electricity|Geothermal'],                
             'IGCC power plant': [],                      
             'light oil combined cycle': [],              
             'gas combined cycle power plant':          ['Secondary Energy|Electricity|Gas|w/o CCS',],        
             'advanced coal power plant with CCS': [],    
             'coal power plant with CCS':               ['Secondary Energy|Electricity|Coal|w/ CCS'],             
             'biomass power plant with CCS':            ['Secondary Energy|Electricity|Biomass|w/ CCS'],          
             'gas combined cycle power plant with CCS': ['Secondary Energy|Electricity|Gas|w/ CCS']
             }
    
# extract results
Energy_Total_R  = np.zeros((len(Regions_R),len(RCPScen_R),len(Industry_R),len(Years_I)))  #riRt
Energy_Mix_R    = np.zeros((len(Regions_R),len(RCPScen_R),len(Industry_R),len(Years_I)))  #riRt
Energy_Total_O  = np.zeros((len(Regions_O),len(RCPScen_R),len(Industry_R),len(Years_I)))  #oiRt
Energy_Mix_O    = np.zeros((len(Regions_O),len(RCPScen_R),len(Industry_R),len(Years_I)))  #oiRt


# Collect energy with RECC classification
for r in range(0,len(Regions_R)):
    for R in range(0,len(RCPScen_R)):
        for i in range(0,len(Industry_R)):
            D = DF.loc[DF['Scenario'].isin(S_DictLookUp[RCPScen_R[R]]) & DF['Region'].isin(R_DictLookUp[Regions_R[r]]) & DF['Model'].isin(Model_I) & DF['Variable'].isin(Ind_DictLookUp[Industry_R[i]]) ] # extract total energy supplied, in EJ/ySL
            D.drop(['Model','Scenario','Region','Variable','Unit'], axis =1, inplace=True)
            Energy_Total_R[r,R,i,:] = D.sum(axis=0)
Energy_Total_O[0,:,:,:] = Energy_Total_R.sum(axis=0)

# From total energy to annual share
for r in range(0,len(Regions_R)):
    for R in range(0,len(RCPScen_R)):
        tot_energy_R = Energy_Total_R[r,R,:,:].sum(axis=0)
        for i in range(0,len(Industry_R)):
            Energy_Mix_R[r,R,i,:] = np.divide(Energy_Total_R[r,R,i,:],tot_energy_R)

for R in range(0,len(RCPScen_R)):
    tot_Energy_O = Energy_Total_O[0,R,:,:].sum(axis=0)
    for i in range(0,len(Industry_R)):
        Energy_Mix_O[0,R,i,:] = np.divide(Energy_Total_O[0,R,i,:],tot_Energy_O)
    
# interpolate
Energy_Mix_R_interp   = np.zeros((len(Regions_R),len(RCPScen_R),len(Industry_R),len(Time_R)))  #riRt
Energy_Mix_O_interp   = np.zeros((len(Regions_O),len(RCPScen_R),len(Industry_R),len(Time_R)))  #oiRt      

for r in range(0,len(Regions_R)):
    for R in range(0,len(RCPScen_R)):
        for i in range(0,len(Industry_R)):    
            f_mix = interp1d(Years_I, Energy_Mix_R[r,R,i,:] , kind='linear')
            Energy_Mix_R_interp[r,R,i,:] = f_mix(Time_R)
            
for R in range(0,len(RCPScen_R)):
    for i in range(0,len(Industry_R)):
        f_mix = interp1d(Years_I, Energy_Mix_O[0,R,i,:] , kind='linear')
        Energy_Mix_O_interp[0,R,i,:] = f_mix(Time_R)
        
        
# export results     
Headers= ['Region','RCP_Scen','Industry']
for t in Time_R:
    Headers.append(str(t))

Results_R = openpyxl.Workbook()
sheet_R = Results_R.active
sheet_R.append(Headers)
Rix = 2
for r in range(0,len(Regions_R)):
    for R in range(0,len(RCPScen_R)):
        for i in range(0,len(Industry_R)):
            sheet_R.cell(Rix,1).value = Regions_R[r]
            sheet_R.cell(Rix,2).value = RCPScen_R[R]
            sheet_R.cell(Rix,3).value = Industry_R[i]
            for t in range(0,len(Time_R)):
                sheet_R.cell(Rix,4+t).value = Energy_Mix_R_interp[r,R,i,t]
            Rix +=1
            
Results_O = openpyxl.Workbook()
sheet_O = Results_O.active
sheet_O.append(Headers)
Rix = 2
for R in range(0,len(RCPScen_R)):
    for i in range(0,len(Industry_R)):
        sheet_O.cell(Rix,1).value = 'World'
        sheet_O.cell(Rix,2).value = RCPScen_R[R]
        sheet_O.cell(Rix,3).value = Industry_R[i]
        for t in range(0,len(Time_R)):
            sheet_O.cell(Rix,4+t).value = Energy_Mix_O_interp[0,R,i,t]
        Rix +=1
        
Results_R.save('3_SHA_ElectricityMix.xlsx')
Results_O.save('3_SHA_ElectricityMix_World.xlsx')

print('Done')
# The end.