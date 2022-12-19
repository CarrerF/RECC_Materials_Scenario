import pandas as pd
import openpyxl 

file = 'ProcessExtensions_Industry.xlsx'

data = pd.read_excel(file, skiprows=[  1])
output = openpyxl.Workbook()
sheet = output.active

Industry_List = ['advanced coal power plant with CCS',
                 'bio powerplant',
                 'biomass power plant with CCS',
                 'coal power plant',
                 'coal power plant with CCS',
                 'coal power plant without abatement measures',
                 'concentrating solar power plant (CSP)',
                 'gas combined cycle power plant',
                 'gas combined cycle power plant with CCS',
                 'geothermal power plant',
                 'hydro power plant',
                 'IGCC power plant',
                 'light oil combined cycle',
                 'nuclear power plant',
                 'oil power plant',
                 'solar photovoltaic power plant',
                 'wind power plant offshore',
                 'wind power plant onshore']

Industry_match = {
     'advanced coal power plant with CCS':          {'industry':None, 'technology':None },
     'bio powerplant':                              {'industry':None, 'technology':None },
     'biomass power plant with CCS':                {'industry':None, 'technology':None },
     'coal power plant':                            {'industry':'Hard coal', 'technology':'PC, without CCS' },
     'coal power plant with CCS':                   {'industry':None, 'technology':None },
     'coal power plant without abatement measures': {'industry':None, 'technology':None },
     'concentrating solar power plant (CSP)':       {'industry':'CSP', 'technology':'tower' },
     'gas combined cycle power plant':              {'industry':'Natural gas', 'technology':'NGCC, without CCS' },
     'gas combined cycle power plant with CCS':     {'industry':'Natural gas', 'technology':'NGCC, with CCS' },
     'geothermal power plant':                      {'industry':None, 'technology':None },
     'hydro power plant':                           {'industry':'Hydro', 'technology':'660 MW' },
     'IGCC power plant':                            {'industry':None, 'technology':None },
     'light oil combined cycle':                    {'industry':None, 'technology':None },
     'nuclear power plant':                         {'industry':'Nuclear', 'technology':'average' },
     'oil power plant':                             {'industry':'Oil', 'technology':'ecoinvent' },
     'solar photovoltaic power plant':              {'industry':'PV', 'technology':'CIGS, roof-mounted' },
     'wind power plant offshore':                   {'industry':'Wind', 'technology':'offshore, steel foundation' },
     'wind power plant onshore':                    {'industry':'Wind', 'technology':'onshore' }
    }

Extensions_List = ['GWP100','Land occupation','Water depletion']

Extensions_match = {
    'GWP100':          'CLIMATE CHANGE TOTAL',
    'Land occupation': 'TOTAL LAND OCCUPATION',
    'Water depletion': 'DISSIPATED WATER'
    }

Headers = ['Industry','Midpoint','region','2015']

sheet.append(Headers)
Rix = 2

for industry in Industry_List:
    for extension in Extensions_List:
        if Industry_match[industry]['industry'] != None:
             dat = float(data[ (data['Industry']==Industry_match[industry]['industry']) & (data['Technology']==Industry_match[industry]['technology']) ][Extensions_match[extension]])
             sheet.cell(Rix,1).value = industry
             sheet.cell(Rix,2).value = extension
             sheet.cell(Rix,3).value = 'World'
             sheet.cell(Rix,4).value = dat
             Rix +=1
        
output.save('4_PE_ProcessExtensions_Industry.xlsx')

print('Done!')
        





