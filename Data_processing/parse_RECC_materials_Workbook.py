# -*- coding: utf-8 -*-
"""
Created on Mon Dec 19 14:46:28 2022

@author: fabic

Script to parse the materials workbook.
Three datasets are delivered:
        - 4_EI_ProcessEnergyIntensity
        - 4_PE_ProcessExtensions_EnergyCarrier
        - 4_PE_ProcessExtensions_Materials
"""

import openpyxl

file = 'RECC_Materials_Workbook_v3.xlsx'

# Load workbook
workbook = openpyxl.load_workbook(file, data_only=True)
Summary = workbook['Summary']

Impacts_list = ['GWP100','Land occupation', 'Water depletion']
Impacts_dict = { # what column in the summary
    'GWP100': 4,
    'Land occupation': 5,
    'Water depletion': 6
    }



# 1) write 4_PE_ProcessExtensions_Materials
MaterialExtension = openpyxl.Workbook()
M_sheet = MaterialExtension.active
Headers = ['Material_Production_i2', 'Env_midpoints', 'SSP_Regions_1','2015' ]
M_sheet.append(Headers)

Rix = 4 # Reading index in Summary sheet
Wix = 2 # Writing index 

while Summary.cell(Rix,3).value != None:
    for impact in Impacts_list:
        M_sheet.cell(Wix,1).value = Summary.cell(Rix,3).value
        M_sheet.cell(Wix,2).value = impact
        M_sheet.cell(Wix,3).value = 'World'
        M_sheet.cell(Wix,4).value = Summary.cell(Rix,Impacts_dict[impact]).value
        Wix +=1
    Rix +=1

MaterialExtension.save('4_PE_ProcessExtensions_Materials.xlsx')    



# 2)  write 4_PE_ProcessExtensions_EnergyCarriers
CarriersExtension = openpyxl.Workbook()
C_sheet = CarriersExtension.active
Headers = ['Energy_Carriers_m6', 'Env_midpoints', 'SSP_Regions_1','2015' ]
C_sheet.append(Headers)

Wix = 2 # Writing index
while Summary.cell(Rix,3).value == None: # Move reading index to first energy carrier
    Rix+=1 

while Summary.cell(Rix,3).value != None:
    for impact in Impacts_list:
        C_sheet.cell(Wix,1).value = Summary.cell(Rix,3).value
        C_sheet.cell(Wix,2).value = impact
        C_sheet.cell(Wix,3).value = 'World'
        C_sheet.cell(Wix,4).value = Summary.cell(Rix,Impacts_dict[impact]).value
        Wix +=1
    Rix +=1

CarriersExtension.save('4_PE_ProcessExtensions_EnergyCarriers.xlsx')   




# 3) write 4_EI_ProcessEnergyIntensity
EnergyIntensity = openpyxl.Workbook()
EI_sheet = EnergyIntensity.active
Headers = ['Material_Production_i2', 'Energy_Carriers_m6', 'Time', 'SSP_Regions_1', 'RCP_Scen','value','unit','Stats_array_string','Comments' ]
EI_sheet.append(Headers)

Wix = 2 # Writing index
Cix = 5 # Colum index 
while Summary.cell(Rix,5).value == None: # Move reading index to first energy carrier
    Rix+=1 
    
while Summary.cell(Rix,Cix).value != None:
    Process_name = Summary.cell(Rix,Cix).value
    while Summary.cell(Rix+1,Cix).value != None:
        EI_sheet.cell(Wix,1).value = Process_name
        EI_sheet.cell(Wix,2).value = Summary.cell(Rix+1,Cix).value
        EI_sheet.cell(Wix,3).value = 2015
        EI_sheet.cell(Wix,4).value = 'World'
        EI_sheet.cell(Wix,5).value = 'Baseline(unmitigated)'
        EI_sheet.cell(Wix,6).value = Summary.cell(Rix+3,Cix).value
        EI_sheet.cell(Wix,7).value = "MJ/kg"
        EI_sheet.cell(Wix,8).value = 'none'
        EI_sheet.cell(Wix,9).value ='none'
        Wix +=1
        Cix +=1
    Cix+=1

EnergyIntensity.save('4_EI_ProcessEnergyIntensity.xlsx')  

    
print('Done!')
        
        
    
