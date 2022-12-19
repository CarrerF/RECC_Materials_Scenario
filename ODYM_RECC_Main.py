# -*- coding: utf-8 -*-
"""
Created on Wed Dec  7 12:02:28 2022

@author: fabic
"""


# Import required libraries:
import os
import sys
import logging as log
import openpyxl
import numpy as np
import time
import datetime
#import scipy.io
import pandas as pd
import shutil   
import uuid
import matplotlib.pyplot as plt   
from matplotlib.lines import Line2D
import importlib
import getpass
from copy import deepcopy
from tqdm import tqdm
import scipy.stats
from scipy.interpolate import interp1d
from scipy.interpolate import make_interp_spline
import pylab
import pickle

import RECC_Paths # Import path file


log.getLogger('matplotlib.font_manager').disabled = True

#import re
__version__ = str('2.5')
##################################
#    Section 1)  Initialize      #
##################################
# add ODYM module directory to system path
sys.path.insert(0, os.path.join(os.path.join(RECC_Paths.odym_path,'odym'),'modules'))
### 1.1.) Read main script parameters
# Mylog.info('### 1.1 - Read main script parameters')
ProjectSpecs_Name_ConFile = 'RECC_Config.xlsx'
Model_Configfile = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,ProjectSpecs_Name_ConFile), data_only=True)
ScriptConfig = {'Model Setting': Model_Configfile['Cover'].cell(4,4).value}
Model_Configsheet = Model_Configfile[ScriptConfig['Model Setting']]
#Read debug modus:   
DebugCounter = 0
while Model_Configsheet.cell(DebugCounter+1, 3).value != 'Logging_Verbosity':
    DebugCounter += 1
ScriptConfig['Logging_Verbosity'] = Model_Configsheet.cell(DebugCounter+1,4).value # Read loggin verbosity once entry was reached.    
# Extract user name from main file
ProjectSpecs_User_Name     = getpass.getuser()

# import packages whose location is now on the system path:    
import ODYM_Classes as msc # import the ODYM class file
importlib.reload(msc)
import ODYM_Functions as msf  # import the ODYM function file
importlib.reload(msf)
import dynamic_stock_model as dsm # import the dynamic stock model library
importlib.reload(dsm)

Name_Script        = Model_Configsheet.cell(6,4).value
if Name_Script != 'ODYM_RECC_Main':  # Name of this script must equal the specified name in the Excel config file
    raise AssertionError('Fatal: The name of the current script does not match to the sript name specfied in the project configuration file. Exiting the script.')
# the model will terminate if the name of the script that is run is not identical to the script name specified in the config file.
Name_Scenario            = Model_Configsheet.cell(7,4).value # Regional scope as torso for scenario name
StartTime                = datetime.datetime.now()
TimeString               = str(StartTime.year) + '_' + str(StartTime.month) + '_' + str(StartTime.day) + '__' + str(StartTime.hour) + '_' + str(StartTime.minute) + '_' + str(StartTime.second)
#DateString               = str(StartTime.year) + '_' + str(StartTime.month) + '_' + str(StartTime.day)
ProjectSpecs_Path_Result = os.path.join(RECC_Paths.results_path, Name_Scenario + '__' + TimeString )

if not os.path.exists(ProjectSpecs_Path_Result): # Create model run results directory.
    os.makedirs(ProjectSpecs_Path_Result)
# Initialize logger
if ScriptConfig['Logging_Verbosity'] == 'DEBUG':
    log_verbosity = eval("log.DEBUG")  
log_filename = Name_Scenario + '__' + TimeString + '.md'
[Mylog, console_log, file_log] = msf.function_logger(log_filename, ProjectSpecs_Path_Result,
                                                      log_verbosity, log_verbosity)
# log header and general information
Time_Start = time.time()
ScriptConfig['Current_UUID'] = str(uuid.uuid4())
Mylog.info('# Simulation from ' + time.asctime())
Mylog.info('Unique ID of scenario run: ' + ScriptConfig['Current_UUID'])

### 1.2) Read model control parameters
Mylog.info('### 1.2 - Read model control parameters')
#Read control and selection parameters into dictionary
ScriptConfig = msf.ParseModelControl(Model_Configsheet,ScriptConfig)

Mylog.info('Script: ' + Name_Script + '.py')
Mylog.info('Model script version: ' + __version__)
Mylog.info('Model functions version: ' + msf.__version__())
Mylog.info('Model classes version: ' + msc.__version__())
Mylog.info('Current User: ' + ProjectSpecs_User_Name)
Mylog.info('Current Scenario: ' + Name_Scenario)
Mylog.info(ScriptConfig['Description'])
Mylog.debug('----\n')

### 1.3) Organize model output folder and logger
Mylog.info('### 1.3 Organize model output folder and logger')
#Copy Config file and model script into that folder
shutil.copy(os.path.join(RECC_Paths.data_path,ProjectSpecs_Name_ConFile), os.path.join(ProjectSpecs_Path_Result, ProjectSpecs_Name_ConFile))
#shutil.copy(Name_Script + '.py'      , os.path.join(ProjectSpecs_Path_Result, Name_Script + '.py'))

#####################################################
#     Section 2) Read classifications and data      #
#####################################################
Mylog.info('## 2 - Read classification items and define all classifications')
### 2.1) # Read model run config data
Mylog.info('### 2.1 - Read model run config data')
# Note: This part reads the items directly from the Exel master,
# will be replaced by reading them from version-managed csv file.
class_filename       = str(ScriptConfig['Version of master classification']) + '.xlsx'
Classfile            = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,class_filename), data_only=True)
Classsheet           = Classfile['MAIN_Table']
MasterClassification = msf.ParseClassificationFile_Main(Classsheet,Mylog)
    
Mylog.info('Read and parse config table, including the model index table, from model config sheet.')
IT_Aspects,IT_Description,IT_Dimension,IT_Classification,IT_Selector,IT_IndexLetter,PL_Names,PL_Description,PL_Version,PL_IndexStructure,PL_IndexMatch,PL_IndexLayer,PrL_Number,PrL_Name,PrL_Comment,PrL_Type,ScriptConfig = msf.ParseConfigFile(Model_Configsheet,ScriptConfig,Mylog)    

Mylog.info('Define model classifications and select items for model classifications according to information provided by config file.')
ModelClassification  = {} # Dict of model classifications
for m in range(0,len(IT_Aspects)):
    ModelClassification[IT_Aspects[m]] = deepcopy(MasterClassification[IT_Classification[m]])
    EvalString = msf.EvalItemSelectString(IT_Selector[m],len(ModelClassification[IT_Aspects[m]].Items))
    if EvalString.find(':') > -1: # range of items is taken
        RangeStart = int(EvalString[0:EvalString.find(':')])
        RangeStop  = int(EvalString[EvalString.find(':')+1::])
        ModelClassification[IT_Aspects[m]].Items = ModelClassification[IT_Aspects[m]].Items[RangeStart:RangeStop]           
    elif EvalString.find('[') > -1: # selected items are taken
        ModelClassification[IT_Aspects[m]].Items = [ModelClassification[IT_Aspects[m]].Items[i] for i in eval(EvalString)]
    elif EvalString == 'all':
        None
    else:
        Mylog.error('Item select error for aspect ' + IT_Aspects[m] + ' were found in datafile.')
        break
    
### 2.2) # Define model index table and parameter dictionary
Mylog.info('### 2.2 - Define model index table and parameter dictionary')
Model_Time_Start = int(min(ModelClassification['Time'].Items))
Model_Time_End   = int(max(ModelClassification['Time'].Items))
Model_Duration   = Model_Time_End - Model_Time_Start + 1

Mylog.info('Define index table dataframe.')
IndexTable = pd.DataFrame({'Aspect'        : IT_Aspects,  # 'Time' and 'Element' must be present!
                            'Description'   : IT_Description,
                            'Dimension'     : IT_Dimension,
                            'Classification': [ModelClassification[Aspect] for Aspect in IT_Aspects],
                            'IndexLetter'   : IT_IndexLetter})  # Unique one letter (upper or lower case) indices to be used later for calculations.

# Default indexing of IndexTable, other indices are produced on the fly
IndexTable.set_index('Aspect', inplace=True)

# Add indexSize to IndexTable:
IndexTable['IndexSize'] = pd.Series([len(IndexTable.Classification[i].Items) for i in range(0, len(IndexTable.IndexLetter))],
                                    index=IndexTable.index)

# list of the classifications used for each indexletter
IndexTable_ClassificationNames = [IndexTable.Classification[i].Name for i in range(0, len(IndexTable.IndexLetter))]

# 2.3) Define shortcuts for the most important index sizes:
Nt = len(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items)
Ne = len(IndexTable.Classification[IndexTable.index.get_loc('Element')].Items)
Nc = len(IndexTable.Classification[IndexTable.index.get_loc('Cohort')].Items)
Nr = len(IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items)
Nl = len(IndexTable.Classification[IndexTable.index.get_loc('Region11')].Items)
No = len(IndexTable.Classification[IndexTable.index.get_loc('Region1')].Items)
NG = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('G')].Items)
Ng = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items)
Np = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('p')].Items)
NB = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items)
NN = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items) # varies: 24 for region-specific nrb and 4 for aggregated global resolution.
NI = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items)
Na = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('a')].Items)
Nx = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('x')].Items)
#NA = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('A')].Items)
NS = len(IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items)
NR = len(IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
Nw = len(IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items)
Nm = len(IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items)
NX = len(IndexTable.Classification[IndexTable.index.get_loc('Extensions')].Items)
Nn = len(IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items)
NV = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('V')].Items)
Ns = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('s')].Items)
#NT = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items)
NL = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items)
NO = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items)    
#IndexTable.loc['t']['Classification'].Items # get classification items

SwitchTime = Nc-Nt+1 # Index of first model year (2016)
# 2.4) Read model data and parameters.
Mylog.info('Read model data and parameters.')

ParFileName = os.path.join(RECC_Paths.data_path,'RECC_ParameterDict_' + ScriptConfig['RegionalScope'] + '.dat')
try: # Load Pickle parameter dict to save processing time
    ParFileObject = open(ParFileName,'rb')  
    ParameterDict = pickle.load(ParFileObject)  
    ParFileObject.close()  
    Mylog.info('Model data and parameters were read from pickled file with pickle file /parameter reading sequence UUID ' + ParameterDict['Checkkey'])
except:
    ParameterDict = {}
    mo_start = 0 # set mo for re-reading a certain parameter
    for mo in range(mo_start,len(PL_Names)):
        #mo = 76 # set mo for re-reading a certain parameter
        #ParPath = os.path.join(os.path.abspath(os.path.join(ProjectSpecs_Path_Main, '.')), 'ODYM_RECC_Database', PL_Version[mo])
        ParPath = os.path.join(RECC_Paths.data_path, PL_Names[mo] + '_' + PL_Version[mo])
        Mylog.info('Reading parameter ' + PL_Names[mo])
        #MetaData, Values = msf.ReadParameter(ParPath = ParPath,ThisPar = PL_Names[mo], ThisParIx = PL_IndexStructure[mo], IndexMatch = PL_IndexMatch[mo], ThisParLayerSel = PL_IndexLayer[mo], MasterClassification,IndexTable,IndexTable_ClassificationNames,ScriptConfig,Mylog) # Do not change order of parameters handed over to function!
        # Do not change order of parameters handed over to function!
        MetaData, Values = msf.ReadParameterXLSX(ParPath, PL_Names[mo], PL_IndexStructure[mo], PL_IndexMatch[mo],
                                              PL_IndexLayer[mo], MasterClassification, IndexTable,
                                              IndexTable_ClassificationNames, ScriptConfig, Mylog, False)
        ParameterDict[PL_Names[mo]] = msc.Parameter(Name=MetaData['Dataset_Name'], ID=MetaData['Dataset_ID'],
                                                    UUID=MetaData['Dataset_UUID'], P_Res=None, MetaData=MetaData,
                                                    Indices=PL_IndexStructure[mo], Values=Values, Uncert=None,
                                                    Unit=MetaData['Dataset_Unit'])
        Mylog.info('Current parameter file UUID: ' + MetaData['Dataset_UUID'])
        Mylog.info('_')
    Mylog.info('Reading of parameters finished.')
    CheckKey = str(uuid.uuid4()) # generate UUID for this parameter reading sequence.
    Mylog.info('Current parameter reading sequence UUID: ' + CheckKey)
    Mylog.info('Entire parameter set stored under this UUID, will be reloaded for future calculations.')
    ParameterDict['Checkkey'] = CheckKey
    # Save to pickle file for next model run
    ParFileObject = open(ParFileName,'wb') 
    pickle.dump(ParameterDict,ParFileObject)   
    ParFileObject.close()
    
Mylog.info('_')
Mylog.info('_')

RECC_System = msc.MFAsystem(Name='RECC_SingleScenario',
                            Geogr_Scope='19 regions + 1 single country', #IndexTableR.Classification[IndexTableR.set_index('IndexLetter').index.get_loc('r')].Items,
                            Unit='Mt',
                            ProcessList=[],
                            FlowDict={},
                            StockDict={},
                            ParameterDict=ParameterDict,
                            Time_Start=Model_Time_Start,
                            Time_End=Model_Time_End,
                            IndexTable=IndexTable,
                            Elements=IndexTable.loc['Element'].Classification.Items,
                            Graphical=None)


#### Start computations of dynamic impact factors ###
ElecIndex        = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('electricity')


# impact of electricity generation for diagnostic
impact_elect =  np.einsum('I,Ix->x',
                    RECC_System.ParameterDict['3_SHA_ElectricityMix_World'].Values[0,0,:,0],              # MJ industry/MJ el        
                    RECC_System.ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,0,0]/3.6,     # impact/MJ industry 
                     )  # Impact/MJ el
impact_elect_alu =  np.einsum('I,Ix->x',
                    RECC_System.ParameterDict['3_SHA_ElectricityMix_World_Alu'].Values[0,0,:,0],              # MJ industry/MJ el        
                    RECC_System.ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,0,0]/3.6,     # impact/MJ industry 
                     )  # Impact/MJ el

# reshape electricity mix: oRit->mi
Par_ElectricityMix_2015  = np.einsum('I,m->mI', RECC_System.ParameterDict['3_SHA_ElectricityMix_World'].Values[0,0,:,0], np.ones(Nm) )
# overwright aluminium energy mix
for mm in [4,5]:
    Par_ElectricityMix_2015[mm,:] = RECC_System.ParameterDict['3_SHA_ElectricityMix_World_Alu'].Values[0,0,:,0]
    
# Fuel contributions
fuel_production = np.einsum('nx,n,Pn->Px',
                     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers'].Values[:,:,0,0], # impact/kg fuel        
                     RECC_System.ParameterDict['3_EI_SpecificEnergy_EnergyCarriers'].Values[:,0],   # kg fuel/MJ fuel 
                     RECC_System.ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,0,0,0],    # MJ fuel/kg mat 
                     )  # Impact/kg mat

# Direct contributions
direct_impact = np.einsum('xn,Pn->Px',
                     RECC_System.ParameterDict['6_PR_DirectEmissions'].Values[:,:],                 # impact/MJ fuel        
                     RECC_System.ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,0,0,0],    # MJ fuel/kg mat 
                     )  # Impact/kg mat

# Electricity generation
elec_production = np.einsum('P,PI,Ix->Px',
                    RECC_System.ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,ElecIndex,0,0,0],    # MJ el/kg mat
                    Par_ElectricityMix_2015,              # MJ industry/MJ el        
                    RECC_System.ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,0,0]/3.6,          # impact/MJ industry 
                     )  # Impact/kg mat
# compute residuals
residuals = RECC_System.ParameterDict['4_PE_ProcessExtensions_Materials'].Values[:,:,0,0] - fuel_production - direct_impact - elec_production   # impact / kg mat

# define new parameter with residual only. This parameter replaces the old 4_PE_ProcessExtension
# msf.Parameter

# later, to get the total impact, sum contribution of:
    # residual (constant over time)
    # +fuel: material_demand*4_EI_ProcessEnergyIntensity*4_PE_ProcessExtensions_EnegryCarriers. ProcessEnergyIntensity in scenario dependent, 4_PE_ProcessExtensions_EnegryCarriers is constanst over time
    # +direct: material_demand*4_EI_ProcessEnergyIntensity*6_PR_DIrect emissions. ProcessEnergyIntensity in scenario dependent, &_DIrectEmissions is constant over time
    # +material_demand*4_EI_ProcessEnergyIntensity[electricity]*3_SHA_ElectricityMix*6_PE_ProcessExtensions_Industry. 3_SHA_ElectricityMix is RCP dependent, 6_PE_ProcessExtensions_Industry is constant over time


##  VISUALIZATION for DIAGNOSTOCS
labels = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items
width = 0.35
fig, ax = plt.subplots()

i=1   # 0:gwp, 1:land, 2:water, 3:material 
ax.bar(labels, fuel_production[:,i], width, label='fuel')
ax.bar(labels, direct_impact[:,i], width,  bottom=fuel_production[:,i], label='direct')
ax.bar(labels, elec_production[:,i], width,  bottom=fuel_production[:,i]+direct_impact[:,i], label='elec')
ax.bar(labels, residuals[:,i], width,  bottom=fuel_production[:,i]+direct_impact[:,i]+elec_production[:,i], label='residual')
ax.set_ylabel('contribution')

ax.set_title('Contributions')
plt.xticks(rotation=90)
ax.legend()
plt.show()


## percentages
tot = fuel_production + direct_impact + elec_production + residuals

prcg_fuel   = np.divide(fuel_production, tot, out=np.zeros_like(fuel_production), where=tot!=0)
prcg_direct = np.divide(direct_impact,   tot, out=np.zeros_like(direct_impact),   where=tot!=0)
prcg_elec   = np.divide(elec_production, tot, out=np.zeros_like(elec_production), where=tot!=0)
prcg_res    = np.divide(residuals,       tot, out=np.zeros_like(residuals),       where=tot!=0)
fig, ax = plt.subplots()
ax.bar(labels, prcg_fuel[:,i],   width, label='fuel')
ax.bar(labels, prcg_direct[:,i], width, bottom=prcg_fuel[:,i], label='direct')
ax.bar(labels, prcg_elec[:,i],   width, bottom=prcg_fuel[:,i] + prcg_direct[:,i], label='elec')
ax.bar(labels, prcg_res[:,i],    width, bottom=prcg_fuel[:,i] + prcg_direct[:,i] + prcg_elec[:,i], label='residual')
ax.set_ylabel('contribution')

ax.set_title('Contributions')
plt.xticks(rotation=90)
ax.legend()
plt.show()

print('Done!')